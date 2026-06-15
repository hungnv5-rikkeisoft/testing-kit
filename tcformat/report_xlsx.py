"""Render Stage 3 report from aggregated YAML results.

Writes the team template's "3. Test Report" sheet plus a new "Evidence" sheet
that embeds per-step screenshots with a caption and a hyperlink to the original.
"""
from __future__ import annotations
import os
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from tcformat.report_sheet import (
    REPORT_SHEET, find_header_row, clear_region, write_screen_row)

EVIDENCE_SHEET = "Evidence"
EVIDENCE_HEADERS = ["TestcaseID", "Browser", "Step", "Anh", "Mo full-size", "Note"]
BROWSERS = ("chrome", "safari")
MAX_IMG_WIDTH = 480  # px


def _sum(reports, browser, key):
    return sum(getattr(r, browser)[key] for r in reports)


def _write_report_sheet(ws, report_data):
    hdr = find_header_row(ws)
    data_start = hdr + 3
    n = len(report_data.screens)
    clear_region(ws, data_start, data_start + max(n, 5) + 10)

    for i, sr in enumerate(report_data.screens):
        write_screen_row(ws, data_start + i, i + 1, sr.screen,
                         sr.executed, sr.chrome, sr.safari, sr.bugs)

    trow = data_start + n
    chrome_tot = {k: _sum(report_data.screens, "chrome", k) for k in ("ok", "ng", "na")}
    safari_tot = {k: _sum(report_data.screens, "safari", k) for k in ("ok", "ng", "na")}
    write_screen_row(ws, trow, None, "Total", report_data.executed,
                     chrome_tot, safari_tot,
                     sum(r.bugs for r in report_data.screens))

    v = trow + 2
    ws.cell(v, 2).value = "Exit criteria"
    ws.cell(v, 3).value = "PASS" if report_data.exit_ok else "FAIL"
    ws.cell(v + 1, 2).value = "Pass rate"
    ws.cell(v + 1, 3).value = f"{report_data.summary.pass_rate:.0%}"
    ws.cell(v + 2, 2).value = "Executed/Planned"
    ws.cell(v + 2, 3).value = f"{report_data.executed}/{report_data.planned}"
    for j, reason in enumerate(report_data.exit_reasons):
        ws.cell(v + 3 + j, 2).value = "Reason"
        ws.cell(v + 3 + j, 3).value = reason


def _embed(ws, row, col, img_abs):
    img = XLImage(str(img_abs))
    if img.width and img.width > MAX_IMG_WIDTH:
        scale = MAX_IMG_WIDTH / img.width
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
    ws.add_image(img, f"{get_column_letter(col)}{row}")
    ws.row_dimensions[row].height = (img.height or 100) * 0.75
    letter = get_column_letter(col)
    cur = ws.column_dimensions[letter].width or 0
    ws.column_dimensions[letter].width = max(cur, (img.width or MAX_IMG_WIDTH) / 7)


def _write_evidence_sheet(wb, screens, base_dir, out_path):
    ws = wb.create_sheet(EVIDENCE_SHEET)
    for col, h in enumerate(EVIDENCE_HEADERS, start=1):
        ws.cell(1, col).value = h
    out_dir = Path(out_path).resolve().parent

    row = 2
    for sc in screens:
        for tc in sc.testcases:
            for br_name in BROWSERS:
                br = getattr(tc.result, br_name)
                note_written = False
                for idx, rel in enumerate(br.evidence, start=1):
                    ws.cell(row, 1).value = tc.id
                    ws.cell(row, 2).value = br_name
                    ws.cell(row, 3).value = f"step_{idx}"
                    img_abs = (Path(base_dir) / rel).resolve()
                    try:
                        link = os.path.relpath(img_abs, out_dir).replace("\\", "/")
                    except ValueError:        # different drive / mount on Windows
                        link = img_abs.as_uri()
                    if img_abs.exists():
                        _embed(ws, row, 4, img_abs)
                    else:
                        ws.cell(row, 4).value = "(file missing)"
                    cell = ws.cell(row, 5)
                    cell.value = "open"
                    cell.hyperlink = link
                    if not note_written and br.note:
                        ws.cell(row, 6).value = br.note
                        note_written = True
                    row += 1
    return ws


def write_report(report_data, screens, template_path, out_path, base_dir=".") -> None:
    wb = load_workbook(template_path)
    _write_report_sheet(wb[REPORT_SHEET], report_data)
    if EVIDENCE_SHEET in wb.sheetnames:
        del wb[EVIDENCE_SHEET]
    _write_evidence_sheet(wb, screens, base_dir, out_path)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
