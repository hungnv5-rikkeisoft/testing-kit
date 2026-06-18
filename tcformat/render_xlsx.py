"""Render Screen objects into the team's testcase-sheet xlsx format.

Clones the template's '4.1.*' sample sheet (preserving styles + summary block),
renames it per screen, and fills the data region from row 10.
"""
from __future__ import annotations
import re
from pathlib import Path
from openpyxl import load_workbook
from tcformat.schema import flatten_expected

DATA_START = 10
LAST_COL = 18  # through column R (Chrome+Safari result columns K..R)
INVALID_TITLE = re.compile(r"[\[\]:\*\?/\\]")

# The template's "1. Record of Change" sheet carries a sample project-name banner
# (merged B14:O14). Overwrite it so the output isn't tied to the template's app.
PROJECT_SHEET = "1. Record of Change"
PROJECT_NAME_CELL = "B14"
DEFAULT_PROJECT_NAME = "Project Name"


def set_project_name(wb, project_name=DEFAULT_PROJECT_NAME) -> None:
    """Set the project-name banner on the Record-of-Change sheet, if present."""
    if PROJECT_SHEET in wb.sheetnames:
        wb[PROJECT_SHEET][PROJECT_NAME_CELL] = project_name


def _find_sample(wb):
    for ws in wb.worksheets:
        if ws.title.strip().startswith("4.1"):
            return ws
    raise ValueError("template missing a '4.1.*' sample testcase sheet")


def _sheet_title(idx: int, name: str) -> str:
    return f"4.{idx} {INVALID_TITLE.sub('', name)}".strip()[:31]


def _clear(ws, first: int, last: int):
    for rng in list(ws.merged_cells.ranges):
        if rng.min_row >= first and rng.max_row <= last:
            ws.unmerge_cells(str(rng))
    for r in range(first, last + 1):
        for c in range(1, LAST_COL + 1):
            ws.cell(r, c).value = None


def _write_result(ws, row, result):
    """Fill result columns K..R from a Result (blank when value is None)."""
    cols = [
        (11, result.chrome.status), (12, result.chrome.bug_id),
        (13, result.chrome.tester), (14, result.chrome.date),
        (15, result.safari.status), (16, result.safari.bug_id),
        (17, result.safari.tester), (18, result.safari.date),
    ]
    for c, v in cols:
        if v is not None:
            ws.cell(row, c).value = v


def _write(ws, testcases, start: int):
    row = start
    no = 0
    last_section = None
    for tc in testcases:
        if tc.section and tc.section != last_section:
            ws.cell(row, 1).value = tc.section
            last_section = tc.section
            row += 1
        no += 1
        ws.cell(row, 1).value = no
        ws.cell(row, 2).value = tc.id
        ws.cell(row, 3).value = tc.main_item
        ws.cell(row, 4).value = tc.middle_item
        ws.cell(row, 5).value = tc.minor_item
        ws.cell(row, 6).value = tc.precondition
        ws.cell(row, 7).value = "\n".join(
            f"{i + 1}. {s}" for i, s in enumerate(tc.steps))
        ws.cell(row, 8).value = "\n".join(
            f"{i + 1}. {flatten_expected(e)}" for i, e in enumerate(tc.expected))
        ws.cell(row, 9).value = tc.type
        ws.cell(row, 10).value = tc.priority
        _write_result(ws, row, tc.result)
        row += 1
    return row


def render_into(wb, screens, project_name=DEFAULT_PROJECT_NAME) -> None:
    """Render screens' testcase sheets into an already-open workbook.

    Clones the template's '4.1.*' sample sheet per screen, fills it, then removes
    the sample. Saving is left to the caller so a report workbook can hold the
    testcase detail, the summary, and evidence in one file.
    """
    set_project_name(wb, project_name)
    sample = _find_sample(wb)
    for idx, screen in enumerate(screens, start=1):
        ws = wb.copy_worksheet(sample)
        ws.title = _sheet_title(idx, screen.screen)
        ws["C1"] = screen.screen
        ws["C2"] = screen.test_level
        # Clear the WHOLE sample data region (the template's sample block runs
        # far below row 10 with formula rows), not just the area we will rewrite,
        # so no ghost sample testcases leak into the output.
        last = max(ws.max_row, DATA_START + len(screen.testcases) * 2 + 5)
        _clear(ws, DATA_START, last)
        _write(ws, screen.testcases, DATA_START)
    wb.remove(sample)


def render(screens, template_path, out_path, project_name=DEFAULT_PROJECT_NAME) -> None:
    wb = load_workbook(template_path)
    render_into(wb, screens, project_name)
    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
