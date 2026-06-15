"""Extract testing objects from the strategy workbook with stable refs.

A ref is "<section>#<stt>", e.g. "2.3.1#1" = sheet section 2.3.1, object STT 1.
Section dividers live in column A as text like "2.3.1 UI testing"; object rows
have an STT number in column A, the object name in column C, the how-to in J.
"""
from __future__ import annotations
import re
from openpyxl import load_workbook

OBJECT_COL = "C"
HOW_COL = "J"
HEADER_TOKEN = "Đối tượng testing"
SECTION_RE = re.compile(r"^(\d+\.\d+\.\d+)")
STRATEGY_SHEETS = ["1_APITesting", "2_IntergrationTesting", "3_System_Testing"]


def list_objects(xlsx_path, sheet_name: str) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    objects: list[dict] = []
    header_seen = False
    current = None
    for row in range(1, ws.max_row + 1):
        a = ws[f"A{row}"].value
        c = ws[f"{OBJECT_COL}{row}"].value
        if c is not None and HEADER_TOKEN in str(c):
            header_seen = True
            continue
        if a is None:
            continue
        a_text = str(a).strip()
        m = SECTION_RE.match(a_text)
        if m and not a_text.replace(".", "").strip().isdigit():
            current = m.group(1)  # section divider like "2.3.1 UI testing"
            continue
        if not header_seen:
            continue
        stt = a_text.split(".")[0]
        if stt.isdigit() and c is not None and str(c).strip():
            how = ws[f"{HOW_COL}{row}"].value
            objects.append({
                "ref": f"{current}#{stt}" if current else None,
                "section": current,
                "stt": stt,
                "object": str(c).strip(),
                "how": str(how).strip() if how else "",
            })
    return objects


def all_refs(xlsx_path) -> set[str]:
    refs: set[str] = set()
    for sheet in STRATEGY_SHEETS:
        for o in list_objects(xlsx_path, sheet):
            if o["ref"]:
                refs.add(o["ref"])
    return refs
