"""Fill the team's "3. Test Report" sheet from pytest JUnit XML output.

Maps a toolkit run (JUnit XML produced by scripts/run.py via --junitxml) into the
company template "Format test case + Test report.xlsx", sheet "3. Test Report".
One row per test file (module). Test status maps to OK (passed) / NG (failed or
error) / N/A (skipped), split across the two browser columns:
  - Google Chrome (window)  <- the desktop (--chrome) JUnit XML
  - Safari (ipad)           <- the tablet  (--safari) JUnit XML, optional

Usage:
    python scripts/gen_report.py --chrome reports/integration-junit.xml \
        --safari reports/integration-tablet-junit.xml \
        --out reports/test_report.xlsx
"""
from __future__ import annotations
import argparse
from pathlib import Path
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from tcformat.report_sheet import (
    REPORT_SHEET, find_header_row, clear_region, write_screen_row)

DEFAULT_TEMPLATE = "template/Format test case + Test report.xlsx"


def parse_junit(path) -> dict[str, dict]:
    """Return {classname: {"ok": int, "ng": int, "na": int}} from a JUnit XML."""
    root = ET.parse(path).getroot()
    suites = root.findall("testsuite") if root.tag == "testsuites" else [root]
    modules: dict[str, dict] = {}
    for suite in suites:
        for tc in suite.findall("testcase"):
            cls = tc.get("classname", "") or tc.get("name", "")
            m = modules.setdefault(cls, {"ok": 0, "ng": 0, "na": 0})
            if tc.find("failure") is not None or tc.find("error") is not None:
                m["ng"] += 1
            elif tc.find("skipped") is not None:
                m["na"] += 1
            else:
                m["ok"] += 1
    return modules


def _screen_name(classname: str) -> str:
    """Display label for a module: the last dotted component (the file stem)."""
    return classname.split(".")[-1] if classname else classname


def build_report(template_path, chrome_junit, safari_junit, out_path) -> dict:
    """Write the Test Report xlsx. Returns the aggregated per-module rows."""
    chrome = parse_junit(chrome_junit) if chrome_junit else {}
    safari = parse_junit(safari_junit) if safari_junit else {}
    modules = sorted(set(chrome) | set(safari))

    wb = load_workbook(template_path)
    ws = wb[REPORT_SHEET]
    hdr = find_header_row(ws)
    data_start = hdr + 3  # header row + 2 sub-header rows, then data

    clear_region(ws, data_start, data_start + max(len(modules), 5) + 5)

    rows = []
    totals = {"c": 4, "total": 0, "c_ok": 0, "c_ng": 0, "c_na": 0,
              "s_ok": 0, "s_ng": 0, "s_na": 0, "bugs": 0}
    for i, cls in enumerate(modules):
        c = chrome.get(cls, {"ok": 0, "ng": 0, "na": 0})
        s = safari.get(cls, {"ok": 0, "ng": 0, "na": 0})
        total = (c["ok"] + c["ng"] + c["na"]) or (s["ok"] + s["ng"] + s["na"])
        bugs = c["ng"] + s["ng"]
        row = data_start + i
        write_screen_row(ws, row, i + 1, _screen_name(cls), total, c, s, bugs)
        rows.append({"screen": _screen_name(cls), "total": total,
                     "chrome": c, "safari": s, "bugs": bugs})
        totals["total"] += total
        totals["c_ok"] += c["ok"]; totals["c_ng"] += c["ng"]; totals["c_na"] += c["na"]
        totals["s_ok"] += s["ok"]; totals["s_ng"] += s["ng"]; totals["s_na"] += s["na"]
        totals["bugs"] += bugs

    trow = data_start + len(modules)
    ws.cell(trow, 2).value = "Total"
    ws.cell(trow, 3).value = totals["total"]
    ws.cell(trow, 4).value = totals["c_ok"]
    ws.cell(trow, 5).value = totals["c_ng"]
    ws.cell(trow, 6).value = totals["c_na"]
    ws.cell(trow, 7).value = totals["s_ok"]
    ws.cell(trow, 8).value = totals["s_ng"]
    ws.cell(trow, 9).value = totals["s_na"]
    ws.cell(trow, 10).value = totals["bugs"]

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return {"rows": rows, "totals": totals}


def build_report_from_yaml(yaml_paths, template_path, out_path, base_dir="."):
    """Stage 3 path: aggregate testcase YAML(s) and write the report workbook.

    Returns the ReportData (caller uses .exit_ok for the process exit code)."""
    from tcformat.schema import load_screen
    from tcformat.report_data import aggregate
    from tcformat.report_xlsx import write_report
    screens = [load_screen(p) for p in yaml_paths]
    data = aggregate(screens)
    write_report(data, screens, template_path, out_path, base_dir=base_dir)
    return data


def main():
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--yaml", action="append", default=None,
                    help="Stage 3: testcase YAML with results (repeatable)")
    ap.add_argument("--chrome", default=None,
                    help="JUnit XML from the desktop/Chrome run (JUnit path)")
    ap.add_argument("--safari", default=None,
                    help="JUnit XML from the iPad/Safari run (optional)")
    ap.add_argument("--template", default=DEFAULT_TEMPLATE)
    ap.add_argument("--out", default="reports/test_report.xlsx")
    args = ap.parse_args()

    if args.yaml:
        data = build_report_from_yaml(args.yaml, args.template, args.out)
        s = data.summary
        print(f"Wrote report -> {args.out} (executed {data.executed}/{data.planned}, "
              f"OK {s.passed} NG {s.failed}, pass {s.pass_rate:.0%}, "
              f"exit {'PASS' if data.exit_ok else 'FAIL'})")
        raise SystemExit(0 if data.exit_ok else 1)

    if not args.chrome:
        ap.error("provide --yaml (Stage 3) or --chrome (JUnit path)")

    result = build_report(args.template, args.chrome, args.safari, args.out)
    t = result["totals"]
    print(f"Wrote {len(result['rows'])} screen row(s) -> {args.out} "
          f"(Chrome OK/NG/N-A {t['c_ok']}/{t['c_ng']}/{t['c_na']}, "
          f"Safari {t['s_ok']}/{t['s_ng']}/{t['s_na']}, bugs {t['bugs']})")


if __name__ == "__main__":
    main()
