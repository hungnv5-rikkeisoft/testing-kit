"""Generate tester checklists from the strategy workbook.

Reads the "Đối tượng testing" (object) + "Cách thức thực hiện và xác nhận" (how)
columns of a strategy sheet and emits a Markdown checklist.

Usage:
    python scripts/gen_checklist.py --xlsx strategy/strategy.xlsx \
        --sheet 1_APITesting --title "API Testing" --out checklists/
"""
from __future__ import annotations
import argparse
from pathlib import Path
from openpyxl import load_workbook

# Column letters that hold the object name / the "how" description per strategy layout.
OBJECT_COL = "C"
HOW_COL = "J"
HEADER_TOKEN = "Đối tượng testing"


def extract_objects(xlsx_path, sheet_name: str) -> list[dict]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    objects: list[dict] = []
    header_seen = False
    for row in range(1, ws.max_row + 1):
        cval = ws[f"{OBJECT_COL}{row}"].value
        if cval is None:
            continue
        text = str(cval).strip()
        if HEADER_TOKEN in text:
            header_seen = True
            continue
        if not header_seen:
            continue
        # A row is a testing object only if column A has an STT number.
        stt = ws[f"A{row}"].value
        if stt is None or not str(stt).strip().split(".")[0].isdigit():
            continue
        how = ws[f"{HOW_COL}{row}"].value
        objects.append({"object": text,
                        "how": str(how).strip() if how else ""})
    return objects


def render_markdown(title: str, objects: list[dict]) -> str:
    lines = [f"# Checklist — {title}", ""]
    for o in objects:
        lines.append(f"- [ ] **{o['object']}**")
        if o["how"]:
            how = o["how"].replace("\n", " ")
            lines.append(f"  - {how}")
    lines.append("")
    return "\n".join(lines)


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--xlsx", default="strategy/strategy.xlsx")
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--title", required=True)
    ap.add_argument("--out", default="checklists/")
    args = ap.parse_args()

    objects = extract_objects(args.xlsx, args.sheet)
    md = render_markdown(args.title, objects)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / f"{args.sheet}.md"
    out_file.write_text(md, encoding="utf-8")
    print(f"Wrote {len(objects)} objects -> {out_file}")


if __name__ == "__main__":
    main()
