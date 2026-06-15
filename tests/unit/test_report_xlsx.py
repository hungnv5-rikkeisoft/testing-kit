from pathlib import Path
from openpyxl import load_workbook
from PIL import Image as PILImage

from tcformat.schema import Screen, Testcase, Result, BrowserResult
from tcformat.report_data import aggregate
from tcformat.report_xlsx import write_report, REPORT_SHEET, EVIDENCE_SHEET

TEMPLATE = "template/Format test case + Test report.xlsx"


def _png(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    PILImage.new("RGB", (600, 400), "white").save(path)


def _screen(evidence_rel):
    return Screen(screen="Basic Info", test_level="IT", testcases=[
        Testcase(id="UI_02", section="UI", main_item="x", type="IT",
                 priority="High",
                 result=Result(chrome=BrowserResult(
                     status="OK", tester="bot", date="2026-06-15",
                     note="looks good", evidence=[evidence_rel]))),
        Testcase(id="FN_02", section="FUNCTION", main_item="y", type="IT",
                 priority="High",
                 result=Result(chrome=BrowserResult(status="NG"))),
    ])


def test_report_sheet_totals_and_exit_block(tmp_path):
    rel = "evidence/basic-info/chrome/UI_02/step_1.png"
    _png(tmp_path / rel)
    sc = _screen(rel)
    out = tmp_path / "reports" / "test_report.xlsx"
    write_report(aggregate([sc]), [sc], TEMPLATE, str(out), base_dir=str(tmp_path))

    ws = load_workbook(out)[REPORT_SHEET]
    from tcformat.report_sheet import find_header_row
    ds = find_header_row(ws) + 3
    assert ws.cell(ds, 2).value == "Basic Info"
    assert ws.cell(ds, 4).value == 1   # chrome OK (UI_02)
    assert ws.cell(ds, 5).value == 1   # chrome NG (FN_02)
    txt = [ws.cell(r, 3).value for r in range(ds, ds + 12)]
    assert "FAIL" in txt          # High-severity NG -> gate fails


def test_evidence_sheet_embeds_image_and_link(tmp_path):
    rel = "evidence/basic-info/chrome/UI_02/step_1.png"
    _png(tmp_path / rel)
    sc = _screen(rel)
    out = tmp_path / "reports" / "test_report.xlsx"
    write_report(aggregate([sc]), [sc], TEMPLATE, str(out), base_dir=str(tmp_path))

    wb = load_workbook(out)
    assert EVIDENCE_SHEET in wb.sheetnames
    ws = wb[EVIDENCE_SHEET]
    assert ws.cell(1, 1).value == "TestcaseID"
    assert ws.cell(2, 1).value == "UI_02"
    assert ws.cell(2, 2).value == "chrome"
    assert ws.cell(2, 3).value == "step_1"
    assert ws.cell(2, 5).hyperlink is not None
    assert ws.cell(2, 6).value == "looks good"
    assert len(ws._images) == 1


def test_missing_evidence_file_does_not_crash(tmp_path):
    rel = "evidence/basic-info/chrome/UI_02/step_1.png"  # NOT created on disk
    sc = _screen(rel)
    out = tmp_path / "reports" / "test_report.xlsx"
    write_report(aggregate([sc]), [sc], TEMPLATE, str(out), base_dir=str(tmp_path))

    ws = load_workbook(out)[EVIDENCE_SHEET]
    assert ws.cell(2, 4).value == "(file missing)"
    assert len(ws._images) == 0


def test_hyperlink_falls_back_when_relpath_raises(tmp_path, monkeypatch):
    import tcformat.report_xlsx as rx
    rel = "evidence/basic-info/chrome/UI_02/step_1.png"
    _png(tmp_path / rel)
    sc = _screen(rel)
    out = tmp_path / "reports" / "test_report.xlsx"

    def boom(*a, **k):
        raise ValueError("different drive")

    monkeypatch.setattr(rx.os.path, "relpath", boom)
    write_report(aggregate([sc]), [sc], TEMPLATE, str(out), base_dir=str(tmp_path))
    ws = load_workbook(out)[EVIDENCE_SHEET]
    assert ws.cell(2, 5).hyperlink is not None
    assert ws.cell(2, 5).hyperlink.target.startswith("file:")


def test_note_written_once_and_safari_evidence_rows(tmp_path):
    rel1 = "evidence/s/safari/FN_01/step_1.png"
    rel2 = "evidence/s/safari/FN_01/step_2.png"
    _png(tmp_path / rel1)
    _png(tmp_path / rel2)
    sc = Screen(screen="S", test_level="IT", testcases=[
        Testcase(id="FN_01", section="FUNCTION", main_item="x", type="IT",
                 priority="High",
                 result=Result(safari=BrowserResult(
                     status="OK", note="safari note", evidence=[rel1, rel2]))),
    ])
    out = tmp_path / "reports" / "r.xlsx"
    write_report(aggregate([sc]), [sc], TEMPLATE, str(out), base_dir=str(tmp_path))
    ws = load_workbook(out)[EVIDENCE_SHEET]
    assert ws.cell(2, 2).value == "safari"
    assert ws.cell(2, 3).value == "step_1"
    assert ws.cell(2, 6).value == "safari note"
    assert ws.cell(3, 3).value == "step_2"
    assert ws.cell(3, 6).value is None        # note only on the first step row


def test_unified_workbook_has_testcase_report_and_evidence(tmp_path):
    rel = "evidence/basic-info/chrome/UI_02/step_1.png"
    _png(tmp_path / rel)
    sc = _screen(rel)
    out = tmp_path / "reports" / "test_report.xlsx"
    write_report(aggregate([sc]), [sc], TEMPLATE, str(out), base_dir=str(tmp_path))

    wb = load_workbook(out)
    # one file holds the summary, the evidence, AND the testcase detail
    assert REPORT_SHEET in wb.sheetnames
    assert EVIDENCE_SHEET in wb.sheetnames
    tc_sheets = [n for n in wb.sheetnames if n.startswith("4.")]
    assert tc_sheets, "expected a 4.x testcase-detail sheet in the workbook"
    ws = wb[tc_sheets[0]]
    ids = {ws.cell(r, 2).value for r in range(10, ws.max_row + 1)}
    assert "UI_02" in ids and "FN_02" in ids   # testcase rows present, same source
    # the unfilled template sample sheet was consumed (no duplicate 4.x sample)
    assert len(tc_sheets) == 1
