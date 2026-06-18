from openpyxl import load_workbook
from tcformat.schema import Screen, Testcase
from tcformat.render_xlsx import render

from tcformat.resources import default_template
TEMPLATE = default_template()


def _sheet_by_c1(wb, name):
    for ws in wb.worksheets:
        if ws["C1"].value == name:
            return ws
    return None


def test_render_fills_testcase_sheet(tmp_path):
    sc = Screen(screen="Login Screen", test_level="IT", testcases=[
        Testcase(id="UI_01", section="UI", main_item="Show",
                 type="IT", priority="High", precondition="none",
                 steps=["Open page", "Click login"], expected=["Form shows"]),
        Testcase(id="FN_01", section="FUNCTION", main_item="Submit",
                 type="IT", priority="Medium",
                 steps=["Fill form"], expected=["Saved"]),
    ])
    out = tmp_path / "login.xlsx"
    render([sc], TEMPLATE, str(out))

    wb = load_workbook(out)
    ws = _sheet_by_c1(wb, "Login Screen")
    assert ws is not None
    assert ws["C2"].value == "IT"
    assert ws["B7"].value == "TestcaseID"  # template header preserved

    # Scan the entire sheet: only the testcase IDs we wrote may appear in
    # column B below the header — no leftover template sample rows (FUNCTION_03..).
    rows = {}
    for r in range(10, ws.max_row + 1):
        b = ws.cell(r, 2).value
        if b:
            rows[b] = r
    assert set(rows) == {"UI_01", "FN_01"}, f"unexpected leftover rows: {set(rows)}"
    r = rows["UI_01"]
    assert "Open page" in (ws.cell(r, 7).value or "")
    assert "Form shows" in (ws.cell(r, 8).value or "")
    assert ws.cell(r, 9).value == "IT"
    assert ws.cell(r, 10).value == "High"


def test_render_flattens_dict_expected(tmp_path):
    sc = Screen(screen="Assert Screen", test_level="IT", testcases=[
        Testcase(id="VAL_01", section="VALIDATION", main_item="Submit",
                 type="IT", priority="High",
                 steps=["Submit empty"],
                 expected=[
                     {"field": "Field A", "value": "XXX"},
                     {"field": "Field B", "enabled": False},
                     {"request": "POST /api/x"},
                 ]),
    ])
    out = tmp_path / "assert.xlsx"
    render([sc], TEMPLATE, str(out))

    wb = load_workbook(out)
    ws = _sheet_by_c1(wb, "Assert Screen")
    cell = ws.cell([r for r in range(10, ws.max_row + 1)
                    if ws.cell(r, 2).value == "VAL_01"][0], 8).value
    assert "1. Field A = XXX" in cell
    assert "2. Field B disabled" in cell
    assert "3. POST /api/x" in cell
