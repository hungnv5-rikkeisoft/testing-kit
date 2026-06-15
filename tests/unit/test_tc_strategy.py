from openpyxl import Workbook
from tcformat.strategy import list_objects, all_refs


def _xlsx(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "2_IntergrationTesting"
    ws["A76"] = "STT"; ws["C76"] = "Đối tượng testing"; ws["J76"] = "Cách thức"
    ws["A77"] = "2.3.1 UI testing"
    ws["A78"] = 1; ws["C78"] = "Giao diện"; ws["J78"] = "Mở màn hình"
    ws["A79"] = 2; ws["C79"] = "Component"; ws["J79"] = "Đếm"
    ws["A88"] = "2.3.2 Functional testing"
    ws["A89"] = 1; ws["C89"] = "Validate"; ws["J89"] = "Nhập"
    p = tmp_path / "s.xlsx"
    wb.save(p)
    return p


def test_list_objects_builds_refs(tmp_path):
    objs = list_objects(_xlsx(tmp_path), "2_IntergrationTesting")
    refs = [o["ref"] for o in objs]
    assert "2.3.1#1" in refs and "2.3.1#2" in refs and "2.3.2#1" in refs
    first = next(o for o in objs if o["ref"] == "2.3.1#1")
    assert first["object"] == "Giao diện"
    assert first["section"] == "2.3.1"


def test_all_refs_on_real_strategy():
    refs = all_refs("strategy/strategy.xlsx")
    assert refs and "2.3.1#1" in refs
