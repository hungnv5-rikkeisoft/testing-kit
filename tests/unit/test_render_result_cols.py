from openpyxl import load_workbook
from tcformat.schema import Screen, Testcase, Result, BrowserResult
from tcformat.render_xlsx import render

from tcformat.resources import default_template
TEMPLATE = default_template()


def _row_of(ws, tc_id):
    for r in range(10, ws.max_row + 1):
        if ws.cell(r, 2).value == tc_id:
            return r
    raise AssertionError(f"{tc_id} not found")


def test_render_writes_result_columns(tmp_path):
    sc = Screen(screen="Result Screen", test_level="IT", testcases=[
        Testcase(id="UI_01", section="UI", main_item="x", type="IT",
                 priority="High", steps=["s"], expected=["e"],
                 result=Result(
                     chrome=BrowserResult(status="OK", tester="bot", date="2026-06-15"),
                     safari=BrowserResult(status="NG", bug_id="BUG-1"))),
        Testcase(id="FN_01", section="FUNCTION", main_item="y", type="IT",
                 priority="Low", steps=["s"], expected=["e"]),  # no result
    ])
    out = tmp_path / "r.xlsx"
    render([sc], TEMPLATE, str(out))
    wb = load_workbook(out)
    ws = next(w for w in wb.worksheets if w["C1"].value == "Result Screen")

    r = _row_of(ws, "UI_01")
    assert ws.cell(r, 11).value == "OK"
    assert ws.cell(r, 13).value == "bot"
    assert ws.cell(r, 14).value == "2026-06-15"
    assert ws.cell(r, 15).value == "NG"
    assert ws.cell(r, 16).value == "BUG-1"

    r2 = _row_of(ws, "FN_01")
    assert ws.cell(r2, 11).value is None
    assert ws.cell(r2, 15).value is None
