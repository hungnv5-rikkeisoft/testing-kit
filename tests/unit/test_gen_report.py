from pathlib import Path
from openpyxl import load_workbook
from scripts.gen_report import parse_junit, build_report

TEMPLATE = "template/Format test case + Test report.xlsx"


def _junit(tmp_path, name, cases):
    """cases: list of (classname, status); status in passed/failed/skipped."""
    parts = []
    for i, (cls, status) in enumerate(cases):
        inner = ""
        if status == "failed":
            inner = "<failure message='x'>boom</failure>"
        elif status == "skipped":
            inner = "<skipped message='s'/>"
        parts.append(
            f"<testcase classname='{cls}' name='t_{i}'>{inner}</testcase>")
    xml = ("<testsuites><testsuite name='pytest' tests='%d'>%s</testsuite>"
           "</testsuites>" % (len(cases), "".join(parts)))
    p = tmp_path / name
    p.write_text(xml, encoding="utf-8")
    return p


def test_parse_junit_counts(tmp_path):
    j = _junit(tmp_path, "c.xml", [
        ("tests.m1", "passed"), ("tests.m1", "failed"), ("tests.m2", "skipped")])
    mods = parse_junit(j)
    assert mods["tests.m1"] == {"ok": 1, "ng": 1, "na": 0}
    assert mods["tests.m2"] == {"ok": 0, "ng": 0, "na": 1}


def _find_header(ws):
    for r in range(1, 20):
        if ws.cell(r, 2).value == "Function/Screen":
            return r
    return None


def test_build_report_fills_sheet(tmp_path):
    chrome = _junit(tmp_path, "chrome.xml", [
        ("tests.integration.test_a", "passed"),
        ("tests.integration.test_a", "failed"),
        ("tests.integration.test_b", "passed")])
    safari = _junit(tmp_path, "safari.xml", [
        ("tests.integration.test_a", "passed"),
        ("tests.integration.test_a", "passed")])
    out = tmp_path / "test_report.xlsx"
    build_report(TEMPLATE, str(chrome), str(safari), str(out))

    ws = load_workbook(out)["3. Test Report"]
    hdr = _find_header(ws)
    assert hdr is not None
    ds = hdr + 3  # data starts 3 rows below the "Function/Screen" header

    # First module row (modules sorted: test_a, test_b)
    assert ws.cell(ds, 2).value == "test_a"
    assert ws.cell(ds, 3).value == 2   # testcase total (chrome)
    assert ws.cell(ds, 4).value == 1   # chrome OK
    assert ws.cell(ds, 5).value == 1   # chrome NG
    assert ws.cell(ds, 6).value == 0   # chrome N/A
    assert ws.cell(ds, 7).value == 2   # safari OK
    assert ws.cell(ds, 10).value == 1  # total bugs (chrome NG 1 + safari NG 0)

    # Total row
    trow = None
    for r in range(ds, ds + 20):
        if ws.cell(r, 2).value == "Total":
            trow = r
            break
    assert trow is not None
    assert ws.cell(trow, 4).value == 2  # chrome OK total (test_a 1 + test_b 1)
    assert ws.cell(trow, 5).value == 1  # chrome NG total
    assert ws.cell(trow, 7).value == 2  # safari OK total


def test_build_report_chrome_only(tmp_path):
    chrome = _junit(tmp_path, "chrome.xml", [("tests.x.test_s", "passed")])
    out = tmp_path / "r.xlsx"
    build_report(TEMPLATE, str(chrome), None, str(out))
    ws = load_workbook(out)["3. Test Report"]
    ds = _find_header(ws) + 3
    assert ws.cell(ds, 2).value == "test_s"
    assert ws.cell(ds, 4).value == 1   # chrome OK
    assert ws.cell(ds, 7).value == 0   # safari OK (no safari junit -> 0)
