from pathlib import Path
from openpyxl import load_workbook
from PIL import Image as PILImage

from tcformat.schema import Screen, Testcase, Result, BrowserResult, dump_screen
from scripts.gen_report import build_report_from_yaml

TEMPLATE = "template/Format test case + Test report.xlsx"


def _png(path):
    path.parent.mkdir(parents=True, exist_ok=True)
    PILImage.new("RGB", (600, 400), "white").save(path)


def test_build_report_from_yaml(tmp_path):
    rel = "evidence/s/chrome/UI_01/step_1.png"
    _png(tmp_path / rel)
    sc = Screen(screen="S", test_level="IT", testcases=[
        Testcase(id="UI_01", section="UI", main_item="x", type="IT",
                 priority="High",
                 result=Result(chrome=BrowserResult(
                     status="OK", evidence=[rel]))),
    ])
    yml = tmp_path / "s.yaml"
    dump_screen(sc, yml)
    out = tmp_path / "reports" / "test_report.xlsx"

    data = build_report_from_yaml([str(yml)], TEMPLATE, str(out),
                                  base_dir=str(tmp_path))
    assert data.executed == 1
    assert data.exit_ok is True   # 100% pass, no bugs
    wb = load_workbook(out)
    assert "3. Test Report" in wb.sheetnames
    assert "Evidence" in wb.sheetnames
