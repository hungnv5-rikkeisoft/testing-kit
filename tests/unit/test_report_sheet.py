from openpyxl import load_workbook
from tcformat.report_sheet import (
    REPORT_SHEET, find_header_row, clear_region, write_screen_row)

TEMPLATE = "template/Format test case + Test report.xlsx"


def test_find_header_row_locates_function_screen():
    ws = load_workbook(TEMPLATE)[REPORT_SHEET]
    assert ws.cell(find_header_row(ws), 2).value == "Function/Screen"


def test_write_screen_row_fills_columns():
    ws = load_workbook(TEMPLATE)[REPORT_SHEET]
    row = find_header_row(ws) + 3
    clear_region(ws, row, row + 5)
    write_screen_row(ws, row, 1, "MyScreen", 3,
                     {"ok": 2, "ng": 1, "na": 0},
                     {"ok": 0, "ng": 0, "na": 0}, 1)
    assert ws.cell(row, 1).value == "1.0"
    assert ws.cell(row, 2).value == "MyScreen"
    assert ws.cell(row, 3).value == 3
    assert ws.cell(row, 4).value == 2   # chrome OK
    assert ws.cell(row, 5).value == 1   # chrome NG
    assert ws.cell(row, 10).value == 1  # bugs


def test_write_screen_row_total_has_no_number():
    ws = load_workbook(TEMPLATE)[REPORT_SHEET]
    row = find_header_row(ws) + 3
    clear_region(ws, row, row + 5)
    write_screen_row(ws, row, None, "Total", 0,
                     {"ok": 0, "ng": 0, "na": 0},
                     {"ok": 0, "ng": 0, "na": 0}, 0)
    assert ws.cell(row, 1).value is None
    assert ws.cell(row, 2).value == "Total"
