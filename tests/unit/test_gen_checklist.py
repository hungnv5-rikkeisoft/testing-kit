from pathlib import Path
from openpyxl import Workbook
from scripts.gen_checklist import extract_objects, render_markdown


def _make_xlsx(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "1_APITesting"
    ws["A72"] = "STT"; ws["C72"] = "Đối tượng testing"
    ws["J72"] = "Cách thức thực hiện và xác nhận"
    ws["A74"] = 1; ws["C74"] = "URL"; ws["J74"] = "Nhập URL và xác nhận domain"
    ws["A79"] = 2; ws["C79"] = "Header"; ws["J79"] = "Nhập header sai/đúng"
    p = tmp_path / "s.xlsx"
    wb.save(p)
    return p


def test_extract_objects_finds_rows(tmp_path):
    xlsx = _make_xlsx(tmp_path)
    objs = extract_objects(xlsx, "1_APITesting")
    names = [o["object"] for o in objs]
    assert "URL" in names and "Header" in names


def test_render_markdown_has_checkboxes(tmp_path):
    objs = [{"object": "URL", "how": "Nhập URL"}]
    md = render_markdown("API Testing", objs)
    assert "- [ ]" in md and "URL" in md
